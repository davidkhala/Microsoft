import csv
from os import PathLike
from pathlib import Path

from openpyxl import load_workbook


def to_csv(source: PathLike, sink: Path, *, keep_formula=True):
    wb = load_workbook(filename=source, read_only=True, data_only=not keep_formula)
    ws = wb.active
    with sink.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
    wb.close()
